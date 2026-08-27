"""主要材料表 / 工程量回写 单测（方案 A）。

覆盖：
- db.summarize_materials：设备按块计数 + 导线按图层长度（×sheet.scale），
  导线层 = linear 桶 ∪ 图层名含关键词（默认 "line" 等）
- boq.write_back_quantities：recompute 结果 → boq_item.measured_qty
- report.export_materials：单 sheet 上下两区（设备+导线）+ 换算率三列 + 告警
- report.export_report：use_measured=True 用实测值、False 用实时计算

不连真实 DB：
- summarize 用内存 SQLite（monkeypatch app.db.get_conn / get_sheets...）
- 回写/导出用 mock 假数据
```
python -m pytest tests/test_materials.py -q
or python -m unittest tests.test_materials -v
```
"""
from __future__ import annotations

import os
import sqlite3
import tempfile
import unittest
from unittest import mock

from openpyxl import load_workbook

import app.db as db
from app.models import BoqItem


def _cm(conn):
    """模拟 get_conn() 的上下文管理器（commit/rollback，不 close）。"""

    class _CM:
        def __enter__(self):
            return conn

        def __exit__(self, *exc):
            if exc[0] is None:
                conn.commit()
            else:
                conn.rollback()
            return False

    return _CM()


_ENTITY_SCHEMA = """
CREATE TABLE entity (
  sheet_id INTEGER, entity_id TEXT, dxf_type TEXT, layer TEXT,
  block_name TEXT, length REAL DEFAULT 0, area REAL DEFAULT 0
);
CREATE TABLE sheet (
  id INTEGER PRIMARY KEY, name TEXT, scale REAL DEFAULT 1.0,
  blocks_json TEXT DEFAULT ''
);
CREATE TABLE project_config (
  project_id INTEGER PRIMARY KEY, layer_rules TEXT DEFAULT '{}'
);
CREATE TABLE boq_item (
  id INTEGER PRIMARY KEY, project_id INTEGER, code TEXT, description TEXT,
  unit TEXT, original_qty REAL DEFAULT 0, rule_type TEXT DEFAULT 'length',
  scale_factor REAL DEFAULT 1.0, measured_qty REAL DEFAULT 0
);
"""


class MaterialsSummaryTest(unittest.TestCase):
    """summarize_materials：内存 SQLite 验证聚合口径"""

    def setUp(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(_ENTITY_SCHEMA)
        # 三张图纸：S1 scale=1.0, S2 scale=2.0, S3 scale=1.0
        conn.executemany("INSERT INTO sheet (id,name,scale) VALUES (?,?,?)",
                         [(1, "S1", 1.0), (2, "S2", 2.0), (3, "S3", 1.0)])
        # 设备：INSERT 块
        conn.executemany("INSERT INTO entity VALUES (?,?,?,?,?,?,?)", [
            (1, "e1", "INSERT", "L-DATA", "FAN", 0, 0),
            (1, "e2", "INSERT", "L-DATA", "FAN", 0, 0),
            (2, "e3", "INSERT", "L-DATA", "FAN", 0, 0),
            (2, "e4", "INSERT", "L-DATA", "LIGHT-BAR", 0, 0),
        ])
        # 导线层（layer 名含线关键词）
        conn.executemany("INSERT INTO entity VALUES (?,?,?,?,?,?,?)", [
            (1, "w1", "LINE", "L-WIRE", "", 10, 0),
            (1, "w2", "LINE", "L-WIRE", "", 5, 0),
            (2, "w3", "LINE", "L-WIRE", "", 3, 0),
            (2, "w4", "POLYLINE", "L-KABLO", "", 4, 0),   # linear 桶命中
            (3, "w5", "LINE", "L-OTHER", "", 999, 0),    # 不命中
        ])
        conn.execute("INSERT INTO project_config (project_id, layer_rules) VALUES (?,?)",
                     (1, '{"layer_rules": {"linear": ["L-KABLO"]}}'))
        self.conn = conn

        p1 = mock.patch.object(db, "get_sheets", side_effect=lambda pid: [
            mock.Mock(id=1, scale=1.0), mock.Mock(id=2, scale=2.0),
            mock.Mock(id=3, scale=1.0)] if pid == 77 else [])
        p1.start(); self.addCleanup(p1.stop)

        p2 = mock.patch.object(db, "get_conn", return_value=_cm(conn))
        p2.start(); self.addCleanup(p2.stop)

        p3 = mock.patch.object(db, "get_project_config", return_value={
            "layer_rules": {"linear": ["L-KABLO"]}})
        p3.start(); self.addCleanup(p3.stop)

    def test_device_count_by_block(self):
        res = db.summarize_materials(77)
        devs = {d["block_name"]: d for d in res["devices"]}
        self.assertEqual(devs["FAN"]["qty"], 3)
        self.assertEqual(devs["FAN"]["sheet_count"], 2)
        self.assertEqual(devs["LIGHT-BAR"]["qty"], 1)
        self.assertEqual(devs["LIGHT-BAR"]["sheet_count"], 1)

    def test_wires_summarized(self):
        res = db.summarize_materials(77)
        w = {x["layer_name"]: x for x in res["wires"]}
        # 长度 = Σ entity.length × sheet.scale
        self.assertAlmostEqual(w["L-WIRE"]["length_raw"], 10 + 5 + 3 * 2)  # 21
        self.assertEqual(w["L-WIRE"]["entity_count"], 3)
        self.assertEqual(w["L-WIRE"]["sheet_count"], 2)
        # linear 桶命中（length 4 × scale 2）
        self.assertAlmostEqual(w["L-KABLO"]["length_raw"], 4 * 2)  # 8
        # 不命中 → 不在结果里
        self.assertNotIn("L-OTHER", w)

    def test_wire_keyword_custom(self):
        res = db.summarize_materials(77, wire_keywords=("KABLO",))
        wires = {x["layer_name"] for x in res["wires"]}
        # 自定义关键词后，只有 "L-KABLO" 命中（"L-WIRE" 不含 "KABLO"）
        self.assertEqual(wires, {"L-KABLO"})

    def test_empty_sheet(self):
        res = db.summarize_materials(999)
        self.assertEqual(res, {"devices": [], "wires": []})


class WriteBackTest(unittest.TestCase):
    """write_back_quantities：回写 recompute 结果到 measured_qty"""

    def test_write_injects_computed_qty(self):
        items = [BoqItem(id=1, project_id=1, code="A", original_qty=0),
                 BoqItem(id=2, project_id=1, code="B", original_qty=0),
                 BoqItem(id=3, project_id=1, code="C", original_qty=0)]
        recomputed = {1: {"qty": 12.5, "count": 3},
                      3: {"qty": 0.0, "count": 0}}
        with mock.patch("app.boq.writeback.db.get_boq_items", return_value=items), \
             mock.patch("app.boq.writeback.recompute", return_value=recomputed), \
             mock.patch("app.boq.writeback.db.update_boq_item") as up:
            from app.boq import write_back_quantities
            res = write_back_quantities(1)
        self.assertEqual(res["total"], 3)
        self.assertEqual(res["written"], 1)  # 只有 qty>0 计 written
        self.assertEqual(up.call_count, 3)
        by_id = {c.args[0]: c.kwargs["measured_qty"] for c in up.call_args_list}
        self.assertEqual(by_id[1], 12.5)
        self.assertEqual(by_id[2], 0.0)

    def test_reset_measured_qty(self):
        conn = mock.MagicMock()
        cur = mock.MagicMock(); cur.rowcount = 4
        conn.execute.return_value = cur
        cm = mock.MagicMock(); cm.__enter__.return_value = conn
        with mock.patch("app.boq.writeback.db.get_conn", return_value=cm):
            from app.boq import reset_measured_qty
            n = reset_measured_qty(7)
        self.assertEqual(n, 4)
        sql, args = conn.execute.call_args[0]
        self.assertIn("measured_qty=0", sql)
        self.assertEqual(args, (7,))


def _materials_data():
    return {
        "devices": [
            {"block_name": "FAN", "qty": 3, "sheet_count": 2, "layer": "L-DATA", "spec": ""},
            {"block_name": "LIGHT", "qty": 1, "sheet_count": 1, "layer": "L-DATA", "spec": ""},
        ],
        "wires": [
            {"layer_name": "L-WIRE", "entity_count": 3, "sheet_count": 2, "length_raw": 26.0},
        ],
    }


class ExportMaterialsTest(unittest.TestCase):
    """report.export_materials：表1导出 + 换算三列 + 告警"""

    def test_export_materials_xlsx(self):
        import app.report as report
        with mock.patch("app.report.db.summarize_materials",
                       return_value=_materials_data()):
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            tmp.close()
            try:
                n = report.export_materials(
                    1, tmp.name,
                    get_spec_fn=lambda kind, key: "SX" if kind == "block" else "SY",
                    rates={"L-WIRE": 2.0})
                self.assertEqual(n, 3)  # 2 设备 + 1 导线
                wb = load_workbook(tmp.name)
            finally:
                os.unlink(tmp.name)
        ws = wb.active
        self.assertEqual(ws.title, "主要材料表")
        # 告警单元格
        warn = [ws.cell(r, c).value for r in range(1, ws.max_row + 1)
                for c in range(1, ws.max_column + 1)
                if "人工换算" in str(ws.cell(r, c).value or "")]
        self.assertTrue(warn)
        # 换算后长度 = raw × rate
        conv_cells = [(ws.cell(r, 4).value, ws.cell(r, 6).value)
                      for r in range(2, ws.max_row + 1)
                      if ws.cell(r, 2).value == "L-WIRE"]
        self.assertEqual(len(conv_cells), 1)
        raw, conv = conv_cells[0]
        self.assertEqual(raw, 26.0)
        self.assertEqual(conv, 52.0)


class ExportReportTest(unittest.TestCase):
    """report.export_report：use_measured 覆盖值 / 默认实时计算"""

    def _items(self):
        return [
            mock.Mock(id=1, code="A-01", description="风机", unit="台",
                      original_qty=10.0, measured_qty=88.0),
            mock.Mock(id=2, code="B-02", description="桥架", unit="m",
                      original_qty=50.0, measured_qty=0.0),
        ]

    def _compute(self, item, sheet_id, ss, ps):
        return {"qty": 5.0, "count": 1, "detail": [], "factor": 1.0}

    def _run(self, use_measured):
        import app.report as report
        with mock.patch("app.report.db.get_boq_items", return_value=self._items()), \
             mock.patch("app.report.measure.compute_item",
                       side_effect=self._compute), \
             mock.patch("app.report.db.get_mappings", return_value=[]):
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            tmp.close()
            try:
                report.export_report(1, 1, tmp.name, use_measured=use_measured)
                wb = load_workbook(tmp.name)
            finally:
                os.unlink(tmp.name)
        ws = wb.active
        return {ws.cell(r, 1).value: ws.cell(r, 4).value   # code -> 图纸计量数量
                for r in range(2, ws.max_row + 1)}

    def test_use_measured_true(self):
        got = self._run(True)
        # A-01 measured=88 → 用 88；B-02 measured=0 → 退回实时 5
        self.assertEqual(got["A-01"], 88.0)
        self.assertEqual(got["B-02"], 5.0)

    def test_use_measured_false(self):
        got = self._run(False)
        self.assertEqual(got["A-01"], 5.0)
        self.assertEqual(got["B-02"], 5.0)


if __name__ == "__main__":
    unittest.main()