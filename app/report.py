"""Excel 报表导出：依据图纸的算量清单（与原 BOQ 对比）"""
from __future__ import annotations

import logging
import time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import db, measure

logger = logging.getLogger(__name__)


def export_report(project_id: int, sheet_id: int, out_path: str,
                  sheet_scale: float = 1.0, project_scale: float = 1.0,
                  use_measured: bool = False) -> int:
    """导出算量清单。返回导出行数

    Args:
        use_measured: True 时「数量」列用 measured_qty（实测数量列，回写值），
                      False 时用当前计算值（实时重算）。
    """
    started = time.perf_counter()
    items = db.get_boq_items(project_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "算量清单"

    headers = ["编号", "描述", "单位", "图纸计量数量", "原清单数量", "差值", "映射方式", "比例因子"]
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    header_font = Font(bold=True)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    red_fill = PatternFill("solid", fgColor="F8CBAD")   # 超出 → 红
    green_fill = PatternFill("solid", fgColor="C6EFCE") # 不足 → 绿

    row = 2
    for item in items:
        result = measure.compute_item(item, sheet_id, sheet_scale, project_scale)
        qty = result["qty"]
        if use_measured:
            mq = getattr(item, "measured_qty", 0) or 0
            qty = mq if mq else qty   # 有实测值用实测，无则退回实时计算
        orig = item.original_qty or 0.0
        diff = round(qty - orig, 4)
        modes = sorted({m.mode for m in db.get_mappings(item.id, sheet_id)})
        mode_str = ",".join(modes) if modes else ""

        vals = [item.code, item.description, item.unit, qty, orig if orig else "", diff,
                mode_str, result["factor"]]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=row, column=c, value=v)

        diff_cell = ws.cell(row=row, column=6)
        if diff > 0:
            diff_cell.fill = red_fill
        elif diff < 0:
            diff_cell.fill = green_fill
        row += 1

    # 列宽
    widths = [14, 48, 8, 16, 14, 12, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    elapsed = (time.perf_counter() - started) * 1000
    logger.info("export_report: project_id=%s sheet_id=%s items=%d elapsed_ms=%.1f", project_id, sheet_id, row - 2, elapsed)
    return row - 2


def export_materials(project_id: int, out_path: str,
                     wire_keywords: tuple = None,
                     get_spec_fn=None,
                     rates: dict = None) -> int:
    """导出「主要材料表」：设备（块计数）+ 导线（图层长度）→ 一张 Excel。

    结构（单 sheet，上下两区）：
        设备区  -> 块名，数量=块引用数（个）
        导线区  -> 图层名，长度=Σ entity.length×sheet.scale（**实际长度，未换算**）
        导线区带「原始长度/换算率/换算后长度」三列，换算率默认 1.0
        （=不换算），UI 可传 rates 覆盖每层换算率；表头红色告警
        「⚠ 原始长度未自动换算，请按图例比例人工换算」。

    Args:
        project_id: 项目 id
        out_path: 输出 xlsx 路径
        wire_keywords: 导线层关键词（None → 默认 "line" 等）
        get_spec_fn: optional (kind, key) -> str（设备/导线规格补全）
        rates: {layer_name: float} 导线换算率（缺省 1.0）
    Returns:
        导出总行数（设备+导线）
    """
    from . import db
    started = time.perf_counter()
    data = db.summarize_materials(project_id, wire_keywords=wire_keywords or ())
    wb = Workbook()
    ws = wb.active
    ws.title = "主要材料表"

    header_fill = PatternFill("solid", fgColor="D9E2F3")
    header_font = Font(bold=True)
    warn_font = Font(bold=True, color="C00000")

    # 设备区
    dev_headers = ["类别", "名称（块名）", "规格/型号", "单位", "数量", "出现图纸", "备注"]
    for c, h in enumerate(dev_headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    for d in data["devices"]:
        spec = get_spec_fn("block", d["block_name"]) if get_spec_fn else ""
        vals = ["设备", d["block_name"], spec, "个", d["qty"], d["sheet_count"], ""]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=row, column=c, value=v)
        row += 1

    # 空一行分隔
    row += 1

    # 导线区
    wire_headers = ["大类", "名称（图层）", "规格/型号", "原始长度", "换算率",
                    "换算后长度", "实体数", "出现图纸", "备注"]
    for c, h in enumerate(wire_headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 换算提醒标记（表头旁）
    cell = ws.cell(row=row, column=len(wire_headers) + 1,
                   value="⚠ 原始长度未自动换算，请按图例比例人工换算")
    cell.font = warn_font
    row += 1

    rates = rates or {}
    for w in data["wires"]:
        spec = get_spec_fn("layer", w["layer_name"]) if get_spec_fn else ""
        rate = rates.get(w["layer_name"], 1.0)
        vals = ["导线", w["layer_name"], spec,
                round(w["length_raw"], 2), rate,
                round(w["length_raw"] * rate, 2),
                w["entity_count"], w["sheet_count"], ""]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=row, column=c, value=v)
        row += 1

    # 列宽
    for i, w_ in enumerate([10, 34, 26, 14, 16, 14, 12, 12, 14, 26], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w_

    wb.save(out_path)
    total = len(data["devices"]) + len(data["wires"])
    elapsed = (time.perf_counter() - started) * 1000
    logger.info("export_materials: project_id=%s devices=%d wires=%d elapsed_ms=%.1f",
                project_id, len(data["devices"]), len(data["wires"]), elapsed)
    return total


def export_items_to_excel(items: list, out_path: str) -> int:
    """导出 AI 算量结果（TakeoffItem 列表）到 Excel。

    22C-2 配套函数：被 ai_results_dialog 调用
    items: list[TakeoffItem]（来自 app.takeoff.orchestrator.TakeoffItem）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "AI 算量结果"

    headers = ["编号", "描述", "单位", "数量", "置信度", "来源", "理由", "冲突"]
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    header_font = Font(bold=True)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 颜色：按置信度 + 冲突
    high_fill = PatternFill("solid", fgColor="C6EFCE")
    mid_fill = PatternFill("solid", fgColor="FFEB9C")
    low_fill = PatternFill("solid", fgColor="FFC7CE")
    conflict_fill = PatternFill("solid", fgColor="F8CBAD")

    row = 2
    for it in items:
        conf = it.confidence
        is_conflict = it.raw.get("_conflict", False)
        vals = [
            it.code, it.description, it.unit, round(it.quantity, 4),
            f"{conf:.0%}", it.source_layer or it.source_block or "-",
            (it.reasoning or "")[:200],
            "是" if is_conflict else "",
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=row, column=c, value=v)
        # 整行染色
        if is_conflict:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = conflict_fill
        elif conf >= 0.7:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = high_fill
        elif conf >= 0.5:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = mid_fill
        else:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = low_fill
        row += 1

    # 列宽
    widths = [14, 40, 8, 12, 10, 24, 50, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    return row - 2
